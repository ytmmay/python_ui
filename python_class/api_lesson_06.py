'''
一、代码做接口自动化测试基本流程
1.接口文档===接口测试用例
2.自动化excel表格读取测试用例数据
3.执行测试===requests去执行===执行结果
4.执行结果与预期结果对比====是否通过===最终结果
5.最终结果写入excel表格
二、封装函数的步骤
1.实现功能代码写出来
2.参数化===可变的数据===形参
3.返回值===函数有没有需要给别人的东西

读取数据：第三方库 openpyxl===excel 表格数据操作 1.读取 2. 写入
excel 表格三大对象
1.文档===加载进来====工作簿对象====load_workbooks====最好放在同一级目录
2.表单对象
3。表格对象
注意：写入生效的话====一定要保存文件===保存之前一定要关闭文件
#加载工作薄
wd=openpyxl.load_workbook("test_case_api.xlsx")
#获取表单
sheet=  wd["register"]
cell=sheet.cell(row=1,column=1)
#获取单元格的值，重新赋值==写入
cell.value="用例编号"
print(cell.value)
wd.save("test_case_api.xlsx")

'''
import requests,pprint
import jsonpath
import  openpyxl
#from openpyxl import  load_workbook

#读取数据的函数
def read_data(filename,sheetname):
    # 加载工作薄
    #wd = openpyxl.load_workbook("test_case_api.xlsx")
    wd = openpyxl.load_workbook(filename)
    # 获取表单
    #sheet = wd["register"]
    sheet = wd[sheetname]
    cases =[]# 装所有用例
    # 自动获取最大行号列号max_column
    max_row = sheet.max_row
    # 取头不取尾+1
    for i in range(2, max_row + 1):
        case1 = dict(
        case_id=sheet.cell(row=i,column=1).value,
        url=sheet.cell(row=i,column=5).value,
        data=sheet.cell(row=i,column=6).value,
        ex_ret=sheet.cell(row=i,column=7).value)
        cases.append(case1)
    return cases
result=read_data("test_case_api.xlsx","register")
#print(result)

#写入结果函数
def write_result(filename,sheetname,final_ret,row,column):
    # 加载工作薄
    wd = openpyxl.load_workbook(filename)
    # 获取表单
    sheet = wd[sheetname]
    # 获取单元格的值，重新赋值==写入
    cell = sheet.cell(row=row, column=column).value = final_ret
    wd.save(filename)
#wd.save("test_case_api.xlsx")


#注册接口===发送接口请求函数
def api_function(url_api,data_api):
    head_register = {"X-Lemonban-Media-Type": "lemonban.v2"}
    response = requests.post(url=url_api, json=data_api, headers=head_register)
    return response.json()

'''
url_api_register = "http://8.129.91.152:8766/futureloan/member/register"
api_data_register = {
        "mobile_phone": "13298568855",
        "pwd": "lemon123456",
        "type": "0",
        "reg_name": "管理员用户lemon"
    }
resp=api_function(url_api_register,api_data_register)
print(resp)
'''
